from openpyxl import load_workbook

workbook = load_workbook(filename="wk.xlsx")

sheet = workbook.active
stack = []
first = False
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=3, values_only=True) :
	if not first:
		first = True
		continue

	a = row[0]
	b = row[1]
	c = row[2]

	if row[0] == None:
		a = ""

	if row[1] == None :
		b = ""
	
	if row[2] == None:
		c = ""

	id_ = a
	st1 = b+"".lower()
	st2 = c+"".lower()

	stack.append({ 'id': id_, 'st1': st1, 'st2': st2, 'val': st1+'-'+st2})
	# print(id_, st1, st2, st1+st2)


def sort_key(e):
  return e['val']

stack.sort(key=sort_key)

def format(stack):
	result = stack[0]['st1'] + ' - ' + stack[0]['st2'] + ' : ' + str(stack[0]['id'])
	for i in range(1, len(stack)):
		if stack[i]['val']==stack[i-1]['val']:
			result = result +', '+ str(stack[i]['id'])
		else :
			result = result + '\n' +  stack[i]['st1'] + ' - ' + stack[i]['st2'] + ' : ' + str(stack[i]['id'])
	return result

result = format(stack)


print(result)

f = open("result.txt", "w")
f.write(result)
f.close()

print('\n\n\nSuccessfully Completed !!\n\n\n')